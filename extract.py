import openpyxl
import re

path = "l1.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
  
# print the total number of rows 
print(sheet_obj.max_row) 
print(sheet_obj.max_column) 

print sheet_obj.cell(row=1, column=1)
print sheet_obj.cell(row=sheet_obj.max_row, column=sheet_obj.max_column).value
arr = []
for i in range(1,sheet_obj.max_row+1):
	temp = []
	for j in range(1,sheet_obj.max_column+1):
		temp.append(sheet_obj.cell(row=i, column=j).value)
	arr.append(temp)

for ar in arr:
	print "# echo " + "\""+ ar[0] + " "+ ar[1]+"\""
	print ar[2]


